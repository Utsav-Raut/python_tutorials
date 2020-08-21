import openpyxl as xl
wb = xl.load_workbook(r'Employee.xlsx')
ws = wb['Emp']
cells = tuple()
for row_num in range(2,ws.max_row + 1):
    if ws.cell(row_num,1).value == 103:
        cells = ws[row_num]
        break

if cells:
    for cell in cells:
        print(cell.value, end = ' ')
else:
    print("the record is not found")