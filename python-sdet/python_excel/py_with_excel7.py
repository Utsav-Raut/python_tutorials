# delete row
import openpyxl as xl
filepath=r'Employee.xlsx'
wb=xl.load_workbook(filepath)
ws=wb['Emp']
for row_num in range(1,ws.max_row+1):
    if ws.cell(row_num,1).value==3:  
        ws.delete_rows(row_num,1)
        wb.save(filepath)
        print('Employee deleted')
        break
else:
    print('Employee not found')

# delete column
import openpyxl as xl
filepath=r'Employee.xlsx'
wb=xl.load_workbook(filepath)
ws=wb['Emp']
for col_num in range(1,ws.max_column+1):
    if ws.cell(1,col_num).value.lower().strip()=='salary':  
        ws.delete_cols(col_num,1)
        wb.save(filepath)
        print('Column deleted')
        break
else:
    print('Column not found')
