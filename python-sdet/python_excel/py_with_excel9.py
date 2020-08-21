import openpyxl as xl
wb = xl.Workbook()
print(wb.sheetnames)
wb.create_sheet(title='Emp', index=0)
print(wb.sheetnames)
wb.remove(wb['Sheet'])
print(wb.sheetnames)
ws=wb['Emp']

ws['A1']="Name"
ws['B1']="salary"
ws.append(["Ramu", 2320.00])
ws.append(["Rakesh", 1951.00])
wb.save(r'Employee1.xlsx')