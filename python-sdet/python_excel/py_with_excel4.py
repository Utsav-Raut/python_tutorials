import openpyxl as xl
wb = xl.load_workbook(r'Employee.xlsx')
ws = wb['Emp']
cells = ws[1]
print('Items of row1 are: ', end = ' ')
for cell in cells:
    print(cell.value, end = ' ')
print()
cells = ws['B']
print('Values of column B are: ')
for cell in cells:
    print(cell.value, end = ' ')