import openpyxl as xl
wb = xl.load_workbook(r'Employee.xlsx')
ws = wb['Emp']
cells = tuple()

for col_num in range(1, ws.max_column + 1):
    if ws.cell(1,col_num).value.lower().strip() == 'name':
        for row_num in range(2, ws.max_row + 1):
            cells += ws.cell(row_num,col_num),

if cells:
    for cell in cells:
        print(cell.value)
else:
    print("no records present")