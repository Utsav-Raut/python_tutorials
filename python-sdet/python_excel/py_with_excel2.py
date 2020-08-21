import openpyxl as xl
wb = xl.load_workbook(r'Employee.xlsx')
ws = wb['Emp']
cell1 = ws.cell(1,1)
# print(cell1)

cell2 = ws['A2']
# print(cell2)

print('Cell1 value = ',cell1.value)
print('Cell2 value = ',cell2.value)
print("No of rows in the worksheet = ",ws.max_row)
print("No of columns in the worksheet = ",ws.max_column)
